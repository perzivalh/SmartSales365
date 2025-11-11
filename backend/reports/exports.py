from __future__ import annotations

import base64
import json
from datetime import datetime
from dataclasses import dataclass
from io import BytesIO
from numbers import Number
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


@dataclass
class ExportedFile:
    filename: str
    content_type: str
    data: bytes

    @property
    def base64(self) -> str:
        return base64.b64encode(self.data).decode("ascii")


def _safe_number(value: object) -> float | None:
    if isinstance(value, Number):
        return float(value)
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _summarize_metrics(columns: Sequence[str], rows: Sequence[dict]) -> list[tuple[str, str]]:
    metrics: list[tuple[str, str]] = [
        ("Filas", str(len(rows))),
        ("Columnas", str(len(columns))),
    ]
    if not rows or not columns:
        return metrics

    numeric_columns = [column for column in columns if any(_safe_number(row.get(column)) is not None for row in rows)]
    if not numeric_columns:
        return metrics

    highlighted = numeric_columns[0]
    numeric_values = [_safe_number(row.get(highlighted)) for row in rows]
    numeric_values = [value for value in numeric_values if value is not None]

    if numeric_values:
        maximum = max(numeric_values)
        total = sum(numeric_values)
        average = total / len(numeric_values)
        formatter = lambda number: f"{number:,.2f}".replace(",", " ").replace(".", ",")
        metrics.extend(
            [
                (f"Máximo ({highlighted})", formatter(maximum)),
                (f"Promedio ({highlighted})", formatter(average)),
                (f"Suma ({highlighted})", formatter(total)),
            ]
        )
    return metrics[:5]


def _looks_like_json(value: object) -> bool:
    if isinstance(value, (dict, list)):
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return (stripped.startswith("{") and stripped.endswith("}")) or (stripped.startswith("[") and stripped.endswith("]"))
    return False


def _select_display_columns(columns: Sequence[str], rows: Sequence[dict], max_columns: int = 6) -> list[str]:
    if len(columns) <= max_columns:
        return list(columns)

    def sample_values(column: str) -> list[object]:
        values: list[object] = []
        for row in rows:
            value = row.get(column)
            if value not in (None, ""):
                values.append(value)
            if len(values) >= 20:
                break
        return values

    def score(column: str) -> float:
        values = sample_values(column)
        if not values:
            avg_length = 0.0
        else:
            avg_length = sum(len(str(value)) for value in values) / len(values)
        numeric_bonus = -120 if values and all(_safe_number(value) is not None for value in values) else 0
        json_penalty = 200 if any(_looks_like_json(value) for value in values) else 0
        name_penalty = 150 if column.lower() in {"metadata", "request_payload", "user_agent"} else 0
        return avg_length + json_penalty + name_penalty + numeric_bonus

    sorted_columns = sorted(columns, key=score)
    return sorted_columns[:max_columns]


def _format_pdf_value(value: object, limit: int = 90) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    else:
        text = str(value)
    text = text.replace("\n", " ").strip()
    if len(text) > limit:
        return text[: limit - 1] + "…"
    return text


def build_excel_export(
    columns: Sequence[str],
    rows: Iterable[dict],
    *,
    prompt: str,
    summary: str,
    generated_at: datetime,
) -> ExportedFile:
    rows_list = list(rows)
    metrics = _summarize_metrics(columns, rows_list)
    display_columns = _select_display_columns(columns, rows_list)

    workbook = Workbook()
    narrative_sheet = workbook.active
    narrative_sheet.title = "Narrativa"
    narrative_sheet.column_dimensions["A"].width = 24
    narrative_sheet.column_dimensions["B"].width = 80

    title_font = Font(size=16, bold=True)
    subtitle_font = Font(size=12, bold=True)
    label_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    narrative_sheet["A1"] = "SmartSales365 - Reporte Inteligente"
    narrative_sheet["A1"].font = title_font

    narrative_sheet["A3"] = "Consulta"
    narrative_sheet["A3"].font = label_font
    narrative_sheet["B3"] = prompt
    narrative_sheet["B3"].alignment = wrap_alignment

    narrative_sheet["A4"] = "Generado"
    narrative_sheet["A4"].font = label_font
    narrative_sheet["B4"] = generated_at.strftime("%d/%m/%Y %H:%M %Z") if getattr(generated_at, "tzinfo", None) else generated_at.strftime("%d/%m/%Y %H:%M")

    narrative_sheet["A6"] = "Resumen ejecutivo"
    narrative_sheet["A6"].font = subtitle_font
    narrative_sheet["B6"] = summary
    narrative_sheet["B6"].alignment = wrap_alignment

    start_row = 8
    if metrics:
        narrative_sheet[f"A{start_row}"] = "Indicadores clave"
        narrative_sheet[f"A{start_row}"].font = subtitle_font
        for offset, (label, value) in enumerate(metrics, start=1):
            label_cell = narrative_sheet.cell(row=start_row + offset, column=1, value=label)
            label_cell.font = label_font
            value_cell = narrative_sheet.cell(row=start_row + offset, column=2, value=value)
            value_cell.alignment = Alignment(vertical="top")
        start_row += len(metrics) + 1

    data_sheet = workbook.create_sheet("Datos")
    data_sheet.append(list(columns))
    for row in rows_list:
        data_sheet.append([row.get(column) for column in columns])

    header_font = Font(bold=True)
    for cell in data_sheet[1]:
        cell.font = header_font

    for index, column in enumerate(columns, start=1):
        candidate_lengths = [len(str(column))]
        for row in rows_list:
            cell_value = row.get(column, "")
            cell_text = "" if cell_value is None else str(cell_value)
            candidate_lengths.append(len(cell_text))
        max_length = max(candidate_lengths) if candidate_lengths else 0
        adjusted_width = min(max_length + 2, 50)
        data_sheet.column_dimensions[get_column_letter(index)].width = adjusted_width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = "reporte_dinamico.xlsx"
    return ExportedFile(
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        data=buffer.getvalue(),
    )


def build_pdf_export(
    columns: Sequence[str],
    rows: Iterable[dict],
    *,
    prompt: str,
    summary: str,
    generated_at: datetime,
) -> ExportedFile:
    rows_list = list(rows)
    metrics = _summarize_metrics(columns, rows_list)
    display_columns = _select_display_columns(columns, rows_list)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.8 * inch,
        rightMargin=0.8 * inch,
        topMargin=0.9 * inch,
        bottomMargin=0.9 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#0b1f3f"),
    )
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1f2937"),
    )
    body_style = ParagraphStyle(
        "BodyText",
        parent=styles["BodyText"],
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#111827"),
    )
    meta_style = ParagraphStyle(
        "Meta",
        parent=styles["BodyText"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#4b5563"),
    )

    generated_text = generated_at.strftime("%d/%m/%Y %H:%M %Z") if getattr(generated_at, "tzinfo", None) else generated_at.strftime("%d/%m/%Y %H:%M")

    story = [
        Paragraph("SmartSales365 · Reporte Inteligente", title_style),
        Spacer(1, 10),
        Paragraph(prompt, section_style),
        Spacer(1, 4),
        Paragraph(f"Generado: {generated_text}", meta_style),
        Spacer(1, 16),
        Paragraph("Resumen ejecutivo", section_style),
        Spacer(1, 6),
    ]

    for paragraph in summary.split("\n"):
        story.append(Paragraph(paragraph.strip() or "-", body_style))
        story.append(Spacer(1, 6))
    story.append(Spacer(1, 8))

    if metrics:
        story.append(Paragraph("Indicadores clave", section_style))
        story.append(Spacer(1, 6))
        metrics_table = Table([[label, value] for label, value in metrics], colWidths=[2.3 * inch, 3.5 * inch])
        metrics_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b1f3f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f3f4f6")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#9ca3af")),
                ]
            )
        )
        story.append(metrics_table)
        story.append(Spacer(1, 16))

    story.append(Paragraph("Detalle tabular", section_style))
    story.append(Spacer(1, 6))
    if len(display_columns) < len(columns):
        story.append(
            Paragraph(
                f"Mostrando {len(display_columns)} de {len(columns)} columnas. Exporta a Excel para ver el detalle completo.",
                meta_style,
            )
        )
        story.append(Spacer(1, 4))

    table_data = [list(display_columns)]
    for row in rows_list:
        formatted_row = []
        for column in display_columns:
            value = row.get(column, "")
            formatted_row.append(_format_pdf_value(value))
        table_data.append(formatted_row)

    if len(table_data) == 1:
        table_data.append(["Sin datos disponibles"] + [""] * (len(display_columns) - 1))

    table = Table(table_data, repeatRows=1, colWidths=None)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )

    story.append(table)
    doc.build(story)
    buffer.seek(0)

    return ExportedFile(
        filename="reporte_dinamico.pdf",
        content_type="application/pdf",
        data=buffer.getvalue(),
    )
